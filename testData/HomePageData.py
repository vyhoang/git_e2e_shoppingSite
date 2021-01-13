import openpyxl


class HomePageData:
    test_homePage_data = [{"firstname": "Mickey", "email": "mickey@gmail.com", "gender": "Male"},
                          {"firstname": "Julia", "email": "julia@gmail.com", "gender": "Female"}]

    # load test data from excel file
    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\TestDocuments\\pythonExcelDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]


